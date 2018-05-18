import openpyxl as op

count = 0
glo = []

wb = op.load_workbook('test.xlsx')
wb1 =op.load_workbook('housing_roster.xlsx')

sheetlist = wb.sheetnames
sheetlist1 = wb1.sheetnames

sheet = wb[sheetlist[0]]
sheet1 = wb1[sheetlist1[0]]


for i in range(2, 3276):
    str1 = sheet1.cell(row=i, column=3).value
    str2 = sheet1.cell(row=i, column=4).value
    name = str1 + "" + str2
    # print("Here")
    for j in range(10, 89072):
        temp = sheet.cell(row=j, column=1).value
        if temp == name:
            glo.append(name)


print(glo)
