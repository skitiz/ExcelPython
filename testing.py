import openpyxl as op

wb = op.load_workbook('Facility Participation Counts.xlsx')
wb.get_sheet_names()

sheet = wb.get_sheet_by_name('Sheet1')

def main():
    percentages = []
    for i in range(21, 52):
        temp = 0
        temp = ((sheet.cell(row = i, column = 3).value * 100))
        temp = temp/(sheet.cell(row = i, column = 1).value)
        sheet.cell(row = i, column = 9).value = temp
        percentages.append(temp)

if __name__ == '__main__':
    main()
