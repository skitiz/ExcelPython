import openpyxl as op
count = 1

facility = ['SRAC', 'RWC', 'GroupEx', 'Climbing Gym', 'Perch', 'Owls Nest']

wb = op.load_workbook('MariettaMemberReport improved.xlsx')
wb1 = op.load_workbook('housing_roster.xlsx')
workbook = op.load_workbook('final_workbook.xlsx')

sheetlist = wb.sheetnames
sheetlist1 = wb1.sheetnames

sheet = wb[sheetlist[0]]
sheet1 = wb1[sheetlist1[0]]

final_list = workbook.sheetnames
final_sheet = workbook[final_list[0]]


def get_facilites(n):
    temp = sheet1.cell(row=n, column=1).value
    final_sheet.cell(row=count, column=2).value = 0
    final_sheet.cell(row=count, column=3).value = 0
    final_sheet.cell(row=count, column=4).value = 0
    final_sheet.cell(row=count, column=5).value = 0
    final_sheet.cell(row=count, column=6).value = 0
    final_sheet.cell(row=count, column=7).value = 0
    while temp in facility:
        if temp == 'SRAC':
            final_sheet.cell(row=count, column=2).value = final_sheet.cell(row=count, column=2).value + sheet1.cell(
                row=n, column=5).value
        if temp == 'RWC':
            final_sheet.cell(row=count, column=3).value = final_sheet.cell(row=count, column=3).value + sheet1.cell(
                row=n, column=5).value
        if temp == 'GroupEx':
            final_sheet.cell(row=count, column=4).value = final_sheet.cell(row=count, column=4).value + sheet1.cell(
                row=n, column=5).value
        if temp == 'Climbing Gym':
            final_sheet.cell(row=count, column=5).value = final_sheet.cell(row=count, column=5).value + sheet1.cell(
                row=n, column=5).value
        if temp == 'Perch':
            final_sheet.cell(row=count, column=6).value = final_sheet.cell(row=count, column=6).value + sheet1.cell(
                row=n, column=5).value
        if temp == 'Owls Nest':
            final_sheet.cell(row=count, column=7).value = final_sheet.cell(row=count, column=7).value + sheet1.cell(
                row=n, column=5).value

        n += 1
        temp = sheet1.cell(row=n, column=1).value

for i range(2, 3276):
    str1 = sheet1.cell(row=i, column=3).value
    str2 = sheet1.cell(row=i, column=3).value
    name = str1 + "" + str2
    for j in range(10, 51757):
        temp = sheet.cell(row=j, column=1)

for i in range(2, 3276):
    str1 = sheet.cell(row=i, column=3).value
    str2 = sheet.cell(row=i, column=4).value
    temp = str1 + " " + str2
    for j in range(10, 51757):
        temp1 = sheet1.cell(row=j, column=1).value
        if temp == temp1:
            final_sheet.cell(row=count, column=1).value = temp1
            get_facilites(j+1)
            count+=1
